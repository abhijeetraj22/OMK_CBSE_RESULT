import tkinter as tk
from tkinter import filedialog, messagebox
import re
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font

# Subject mappings
GRADE12_SUBJECTS = {
    '030': 'ECONOMICS',
    '041': 'MATHEMATICS',
    '042': 'PHYSICS',
    '043': 'CHEMISTRY',
    '044': 'BIOLOGY',
    '048': 'PHYSICAL EDUCATION',
    '054': 'BUSINESS STUDIES',
    '055': 'ACCOUNTANCY',
    '065': 'INFORMATICS PRAC.',
    '301': 'ENGLISH CORE'
}

GRADE10_SUBJECTS = {
    '041': 'MATHEMATICS STANDARD',
    '085': 'HINDI COURSE-B',
    '086': 'SCIENCE',
    '087': 'SOCIAL SCIENCE',
    '122': 'SANSKRIT',
    '184': 'ENGLISH LNG & LIT.',
    '241': 'MATHEMATICS BASIC',
    '402': 'INFORMATION TECHNOLOGY'
}

MAIN_SUBJECTS_10 = {'MATHEMATICS STANDARD', 'MATHEMATICS BASIC', 'SCIENCE', 'SOCIAL SCIENCE', 'ENGLISH LNG & LIT.', 'HINDI COURSE-B', 'SANSKRIT'}
MAIN_SUBJECTS_12_PCM = {'ENGLISH CORE', 'PHYSICS', 'CHEMISTRY', 'MATHEMATICS', 'PHYSICAL EDUCATION', 'INFORMATICS PRAC.'}
MAIN_SUBJECTS_12_PCB = {'ENGLISH CORE', 'PHYSICS', 'CHEMISTRY', 'BIOLOGY', 'PHYSICAL EDUCATION', 'INFORMATICS PRAC.'}
MAIN_SUBJECTS_12_COM = {'ENGLISH CORE', 'ECONOMICS', 'BUSINESS STUDIES', 'ACCOUNTANCY', 'PHYSICAL EDUCATION', 'INFORMATICS PRAC.'}


class CBSEParserApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CBSE Result")
        self.resizable(False, False)
        width, height = 600, 400
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
        self.file_path = None
        self.sample_line = ""
        self.grade12 = False
        self.subject_count = 6
        self.init_frame1()
                # Load icon safely (for PyInstaller compatibility)
        def resource_path(relative_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, relative_path)

        try:
            self.iconbitmap(resource_path("cbse_TNC_icon.ico"))
        except Exception as e:
            print("Icon load failed:", e)

    def init_frame1(self):
        self.frame1 = tk.Frame(self)
        self.frame1.pack(fill='both', expand=True)
        tk.Label(self.frame1, text="Step 1: Upload TXT File", font=("Arial", 14)).pack(pady=10)
        tk.Button(self.frame1, text="Choose File", command=self.select_file, font=("Tahoma", 12), width=15, bd=0, bg="pink").pack(pady=5)
        self.file_label = tk.Label(self.frame1, text="No file selected", font=("Arial", 10))
        self.file_label.pack(pady=5)
        tk.Label(self.frame1, text="Step 2: Paste Sample Student Line", font=("Arial", 14)).pack(pady=10)
        self.sample_entry = tk.Text(self.frame1, height=4, width=70)
        self.sample_entry.pack(pady=5)
        tk.Button(self.frame1, text="Next", command=self.handle_sample, font=("Tahoma", 12), width=15, bd=0, bg="lightgreen").pack(pady=10)

    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
        if path:
            self.file_path = path
            self.file_label.config(text=os.path.basename(path))

    def handle_sample(self):
        self.sample_line = self.sample_entry.get("1.0", tk.END).strip()
        if not self.file_path or not self.sample_line:
            messagebox.showerror("Error", "Please select a file and enter a sample line.")
            return
        self.grade12 = bool(re.search(r'(A[1-2]|B[1-2]|C[1-2]|D[1-2])\s+(PASS|FAIL|COMP|ABST)', self.sample_line))
        self.subject_count = 5 if self.grade12 else 6
        self.frame1.destroy()
        self.init_frame2()

    def init_frame2(self):
        self.frame2 = tk.Frame(self)
        self.frame2.pack(fill='both', expand=True)
        tk.Label(self.frame2, text="Step 3: Save Excel Output", font=("Arial", 14)).pack(pady=40)
        tk.Button(self.frame2, text="Save and Generate Excel", command=self.generate_excel, font=("Tahoma", 12), width=25, bd=0, bg="lightblue").pack(pady=20)

    def generate_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            return
        try:
            self.parse_and_save(file_path)
            response = messagebox.askyesno("Success", f"\u2705 File saved:\n{file_path}\n\nGenerate another file?")
            if response:
                self.frame2.destroy()
                self.init_frame1()
            else:
                self.frame2.destroy()
                self.init_final_frame()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def parse_and_save(self, filename):
        with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        data = []
        skip_patterns = [r'^DATE:-', r'^ROLL\s+F', r'^NO\s+L', r'^-+$', r'^SCHOOL\s+:\s+-', r'^TOTAL CANDIDATES', r'^\f', r'^$', r'^REGION:']
        i = 0

        while i < len(lines) - 1:
            line1 = lines[i].strip()
            line2 = lines[i + 1].strip()
            if any(re.match(p, line1) for p in skip_patterns) or not re.match(r'^\d{8}', line1):
                i += 1
                continue

            match = re.match(r'^(\d{8})\s+([MF])\s+(.*?)\s+(\d{3})', line1)
            if not match:
                i += 2
                continue

            roll_no, gender, name = match.group(1), match.group(2), match.group(3).strip()
            subject_codes = re.findall(r'\b\d{3}\b', line1)[:self.subject_count]
            marks_grades = re.findall(r'(\d{2,3})\s+([A-D][1-2])', line2)

            if len(subject_codes) < self.subject_count or len(marks_grades) < self.subject_count:
                i += 2
                continue

            result_match = re.search(r'(PASS|FAIL|COMP|ABST)', line1)
            result = result_match.group(1) if result_match else ""
            subject_map = GRADE12_SUBJECTS if self.grade12 else GRADE10_SUBJECTS
            subjects = [subject_map.get(code, f"Sub{idx+1}") for idx, code in enumerate(subject_codes)]

            row = {'Roll No': roll_no, 'Gender': gender, 'Name': name}
            scores = []

            for idx, subject in enumerate(subjects):
                mark, grade = marks_grades[idx]
                row[subject + " Grade"] = grade
                row[subject] = int(mark)
                scores.append((subject, int(mark)))

            if self.grade12:
                total = sum(m for _, m in scores)
                percentage = round(total / len(scores), 2)
                row.update({'Result': result, 'Total': total, 'Percentage': percentage})
            else:
                main_pool = MAIN_SUBJECTS_10
                main_scores = [m for s, m in scores if s in main_pool]
                top5_scores = sorted([m for _, m in scores], reverse=True)[:5]
                row.update({
                    'Result': result,
                    'Main Total': sum(main_scores) if main_scores else '',
                    'Main %': round(sum(main_scores) / len(main_scores), 2) if main_scores else '',
                    'Top 5 Total': sum(top5_scores),
                    'Top 5 %': round(sum(top5_scores) / 5, 2)
                })

            data.append(row)
            i += 2

        df = pd.DataFrame(data)

        # Reorder columns
        all_subjects = sorted(set(k.replace(" Grade", "") for row in data for k in row if " Grade" in k))
        grade_columns = [s + " Grade" for s in all_subjects]
        mark_columns = all_subjects
        base_columns = ['Roll No', 'Gender', 'Name'] + grade_columns + mark_columns + ['Result']

        if self.grade12:
            final_columns = base_columns + ['Total', 'Percentage']
        else:
            final_columns = base_columns + ['Main Total', 'Main %', 'Top 5 Total', 'Top 5 %']

        df = df.reindex(columns=final_columns)
        df.to_excel(filename, index=False)

        # ----- Subject Summary Section -----
        wb = load_workbook(filename)
        ws = wb.active
        subjects = [col for col in df.columns if col not in ['Roll No', 'Gender', 'Name', 'Result', 'Total', 'Percentage', 'Main Total', 'Main %', 'Top 5 Total', 'Top 5 %'] and not col.endswith("Grade")]

        summary = {
            "Highest in Subject": [df[s].max() for s in subjects],
            "Lowest in Subject": [df[s].min() for s in subjects],
            "Average in Subject": [round(df[s].mean(), 2) for s in subjects],
            "Distinction in Subject": [(df[s] >= 75).sum() for s in subjects],
            "Distinction in Percent": [round((df[s] >= 75).sum() * 100 / len(df), 2) for s in subjects],
            "Obtained 100 out of 100": [(df[s] == 100).sum() for s in subjects],
        }

        start_row = ws.max_row + 2
        ws.cell(row=start_row, column=1).value = "Name of Candidates"
        for col_index, subject in enumerate(subjects, start=2):
            ws.cell(row=start_row, column=col_index).value = subject

        for i, (label, values) in enumerate(summary.items(), start=start_row + 1):
            ws.cell(row=i, column=1).value = label
            for j, value in enumerate(values, start=2):
                ws.cell(row=i, column=j).value = value
                if label == "Highest in Subject":
                    ws.cell(row=i, column=j).font = Font(color="FF0000")
                elif label in ["Average in Subject", "Distinction in Subject"]:
                    ws.cell(row=i, column=j).font = Font(color="0000FF")
            ws.cell(row=i, column=1).font = Font(bold=True)

        wb.save(filename)

    def init_final_frame(self):
        self.final_frame = tk.Frame(self)
        self.final_frame.pack(fill='both', expand=True)
        tk.Label(self.final_frame, text="\ud83c\udf89 Thank you for using the CBSE Result Parser!", font=("Arial", 16), fg='green').pack(pady=60)
        tk.Button(self.final_frame, text="Exit", command=self.destroy, font=("Tahoma", 12), width=20, bd=0, bg="red", fg="white").pack(pady=10)


if __name__ == "__main__":
    app = CBSEParserApp()
    app.mainloop()
