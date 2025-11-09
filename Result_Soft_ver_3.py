import tkinter as tk
from tkinter import filedialog, messagebox
import re
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font



# Grade 10 and 12 subject header mappings
GRADE10_SUBJECT_HEADER_MAP = {
    "184": ("ENG-184", "184-G1"),
    "085": ("HND-085", "085-G2"),
    "122": ("SNK-122", "122-G3"),
    "041": ("MAT-S-041", "041-G4"),
    "241": ("MAT-B-241", "241-G5"),
    "086": ("SCI-086", "086-G6"),
    "087": ("SST-087", "087-G7"),
    "402": ("IT-402", "402-G8")
}

GRADE12_SUBJECT_HEADER_MAP = {
    "301": ("ENG-301", "301-G1"),
    "042": ("PHY-042", "042-G2"),
    "043": ("CHE-043", "043-G3"),
    "044": ("BIO-044", "044-G4"),
    "041": ("MAT-041", "041-G5"),
    "055": ("ACC-055", "055-G6"),
    "054": ("BST-054", "054-G7"),
    "030": ("ECO-030", "030-G8"),
    "048": ("PHED-048", "048-G9"),
    "065": ("IP-065", "065-G10")
}
MAIN_SUBJECTS_10 = {
    "MAT-S-041",  # MATHEMATICS STANDARD
    "MAT-B-241",  # MATHEMATICS BASIC
    "SCI-086",    # SCIENCE
    "SST-087",    # SOCIAL SCIENCE
    "ENG-184",    # ENGLISH LNG & LIT.
    "HND-085",    # HINDI COURSE-B
    "SNK-122"     # SANSKRIT
}

MAIN_SUBJECTS_12_PCM = {
    "ENG-301",    # ENGLISH CORE
    "PHY-042",    # PHYSICS
    "CHE-043",    # CHEMISTRY
    "MAT-041",    # MATHEMATICS
    "PHED-048",   # PHYSICAL EDUCATION
    "IP-065"      # INFORMATICS PRAC.
}

MAIN_SUBJECTS_12_PCB = {
    "ENG-301",    # ENGLISH CORE
    "PHY-042",    # PHYSICS
    "CHE-043",    # CHEMISTRY
    "BIO-044",    # BIOLOGY
    "PHED-048",   # PHYSICAL EDUCATION
    "IP-065"      # INFORMATICS PRAC.
}

MAIN_SUBJECTS_12_COM = {
    "ENG-301",    # ENGLISH CORE
    "ECO-030",    # ECONOMICS
    "BST-054",    # BUSINESS STUDIES
    "ACC-055",    # ACCOUNTANCY
    "PHED-048",   # PHYSICAL EDUCATION
    "IP-065"      # INFORMATICS PRAC.
}



# Utility to generate headers based on subject code
def generate_custom_headers(codes, is_grade12):
    header_map = GRADE12_SUBJECT_HEADER_MAP if is_grade12 else GRADE10_SUBJECT_HEADER_MAP
    result = {}
    for idx, code in enumerate(codes):
        marks_header, grade_header = header_map.get(code, (f"Sub{idx+1}-{code}", f"{code}-G{idx+1}"))
        result[code] = (marks_header, grade_header)
    return result

class CBSEParserApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CBSE Result Soft 1.3")
        self.geometry("600x400")
        self.resizable(False, False)
        width, height = 600, 400
        screen_width = self.winfo_screenwidth()
        screen_height = self.winfo_screenheight()
        x = (screen_width // 2) - (width // 2)
        y = (screen_height // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")

        self.file_path = None
        self.sample_line = ""
        self.grade12 = False
        self.subject_count = 6

        self.init_frame1()
        # Load icon safely (for PyInstaller compatibility)
        def resource_path(relative_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, relative_path)

        try:
            self.iconbitmap(resource_path("cbse_TNC_icon.ico"))
        except Exception as e:
            print("Icon load failed:", e)


    def init_frame1(self):
        self.frame1 = tk.Frame(self)
        self.frame1.pack(fill='both', expand=True)

        tk.Label(self.frame1, text="Step 1: Upload TXT File", font=("Arial", 14)).pack(pady=10)
        tk.Button(self.frame1, text="Choose File", command=self.select_file,
                  font=("Tahoma", 12), width=15, bd=0, bg="pink").pack(pady=5)

        self.file_label = tk.Label(self.frame1, text="No file selected", font=("Arial", 10))
        self.file_label.pack(pady=5)

        tk.Label(self.frame1, text="Step 2: Paste Sample Student Line", font=("Arial", 14)).pack(pady=10)
        self.sample_entry = tk.Text(self.frame1, height=4, width=70)
        self.sample_entry.pack(pady=5)

        tk.Button(self.frame1, text="Next", command=self.handle_sample,
                  font=("Tahoma", 12), width=15, bd=0, bg="lightgreen").pack(pady=10)

    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
        if path:
            self.file_path = path
            self.file_label.config(text=os.path.basename(path))

    def handle_sample(self):
        self.sample_line = self.sample_entry.get("1.0", tk.END).strip()
        if not self.file_path or not self.sample_line:
            messagebox.showerror("Error", "Please select a file and enter a sample line.")
            return

        self.grade12 = bool(re.search(r'(A[1-2]|B[1-2]|C[1-2]|D[1-2])\s+(PASS|FAIL|COMP|ABST)', self.sample_line))
        self.subject_count = 5 if self.grade12 else 6
        self.frame1.destroy()
        self.init_frame2()

    def init_frame2(self):
        self.frame2 = tk.Frame(self)
        self.frame2.pack(fill='both', expand=True)
        tk.Label(self.frame2, text="Step 3: Save Excel Output", font=("Arial", 14)).pack(pady=40)
        tk.Button(self.frame2, text="Save and Generate Excel", command=self.generate_excel,
                  font=("Tahoma", 12), width=25, bd=0, bg="lightblue").pack(pady=20)

    def generate_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            return
        try:
            self.parse_and_save(file_path)
            response = messagebox.askyesno("Success", f"\u2705 File saved:\n{file_path}\n\nGenerate another file?")
            if response:
                self.frame2.destroy()
                self.init_frame1()
            else:
                self.frame2.destroy()
                self.init_final_frame()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # Inside parse_and_save (replace your entire existing method with this):
    def parse_and_save(self, filename):
        with open(self.file_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()

        data = []
        skip_patterns = [r'^DATE:-', r'^ROLL\s+F', r'^NO\s+L', r'^-+$',
                        r'^SCHOOL\s+:\s+-', r'^TOTAL CANDIDATES', r'^\f', r'^$', r'^REGION:']
        i = 0

        while i < len(lines) - 1:
            line1 = lines[i].strip()
            line2 = lines[i + 1].strip()
            if any(re.match(p, line1) for p in skip_patterns) or not re.match(r'^\d{8}', line1):
                i += 1
                continue

            match = re.match(r'^(\d{8})\s+([MF])\s+(.*?)\s+(\d{3})', line1)
            if not match:
                i += 2
                continue

            roll_no, gender, name = match.group(1), match.group(2), match.group(3).strip()
            subject_codes = re.findall(r'\b\d{3}\b', line1)
            marks_grades = re.findall(r'(\d{2,3})\s+([A-D][1-2])', line2)

            if len(subject_codes) < 5 or len(marks_grades) < 5:
                i += 2
                continue

            result_match = re.search(r'(PASS|FAIL|COMP|ABST)', line1)
            result = result_match.group(1) if result_match else ""

            # Decide subject map
            header_map = GRADE12_SUBJECT_HEADER_MAP if self.grade12 else GRADE10_SUBJECT_HEADER_MAP
            row = {'Roll No': roll_no, 'Gender': gender, 'Name': name}
            scores = []

            for idx, code in enumerate(subject_codes[:len(marks_grades)]):
                mark, grade = marks_grades[idx]
                label, grade_label = header_map.get(code, (f"SUB-{code}", f"{code}-G{idx+1}"))
                row[grade_label] = grade
                row[label] = int(mark)
                scores.append((label, int(mark), grade))

            # Grade 12 stream detection and Main % calc
            if self.grade12:
                all_subjects = {s for s, _, _ in scores}
                if all_subjects & MAIN_SUBJECTS_12_PCM:
                    main_set = MAIN_SUBJECTS_12_PCM
                elif all_subjects & MAIN_SUBJECTS_12_PCB:
                    main_set = MAIN_SUBJECTS_12_PCB
                else:
                    main_set = MAIN_SUBJECTS_12_COM

                main_scores = [m for s, m, _ in scores if s in main_set]
                total = sum([m for _, m, _ in scores])
                row.update({
                    'Result': result,
                    'Total': total,
                    'Main %': round(sum(main_scores) / len(main_scores), 2) if main_scores else '',
                    'Main % Rank': ''
                })
            else:
                # Grade 10 calculation
                main_scores = [m for s, m, _ in scores if s in MAIN_SUBJECTS_10]
                top5 = sorted([m for _, m, _ in scores], reverse=True)[:5]
                row.update({
                    'Result': result,
                    'Main Total': sum(main_scores),
                    'Main %': round(sum(main_scores) / len(main_scores), 2) if main_scores else '',
                    'Main % Rank': '',
                    'Top 5 Total': sum(top5),
                    'Top 5 %': round(sum(top5) / 5, 2),
                    'Top 5 % Rank': ''
                })

            data.append(row)
            i += 2

        # Create DataFrame and rank columns
        df = pd.DataFrame(data)

        if 'Main %' in df.columns:
            df['Main % Rank'] = df['Main %'].rank(method='dense', ascending=False).astype(int)
        if 'Top 5 %' in df.columns:
            df['Top 5 % Rank'] = df['Top 5 %'].rank(method='dense', ascending=False).astype(int)

        # Custom column order for output
        fixed_cols = ['Roll No', 'Gender', 'Name']

        if self.grade12:
            ordered_grade_cols = [
                'ENG-301-G', 'PHY-042-G', 'CHE-043-G', 'BIO-044-G', 'MAT-041-G',
                'ACC-054-G', 'BST-055-G', 'ECO-030-G', 'PHED-048-G', 'IP--065-G'
            ]
            ordered_mark_cols = [
                'ENG-301', 'PHY-042', 'CHE-043', 'BIO-044', 'MAT-041',
                'ACC-054', 'BST-055', 'ECO-030', 'PHED-048', 'IP--065'
            ]
            result_cols = ['Result', 'Total', 'Percentage', 'Main %', 'Main % Rank']
        else:
            ordered_grade_cols = [
                'ENG-184-G', 'HND-085-G', 'SNK-122-G', 'MAT-041-G', 'MAT-241-G',
                'SCI-086-G', 'SST-087-G', 'IT--402-G'
            ]
            ordered_mark_cols = [
                'ENG-184', 'HND-085', 'SNK-122', 'MAT-041', 'MAT-241',
                'SCI-086', 'SST-087', 'IT--402'
            ]
            result_cols = [
                'Result', 'Main Total', 'Main %', 'Main % Rank',
                'Top 5 Total', 'Top 5 %', 'Top 5 % Rank'
            ]

        # Filter to only existing columns
        ordered_grade_cols = [col for col in ordered_grade_cols if col in df.columns]
        ordered_mark_cols = [col for col in ordered_mark_cols if col in df.columns]
        result_cols = [col for col in result_cols if col in df.columns]

        df = df[fixed_cols + ordered_grade_cols + ordered_mark_cols + result_cols]

        df.to_excel(filename, index=False)

        # Summary Sheet
        wb = load_workbook(filename)
        ws = wb.active
        summary_subjects = ordered_mark_cols
        summary = {
            "Subject": summary_subjects,
            "Highest in Subject": [df[s].max() for s in summary_subjects],
            "Lowest in Subject": [df[s].min() for s in summary_subjects],
            "Average in Subject": [round(df[s].mean(), 2) for s in summary_subjects],
            "Distinction in Subject": [(df[s] >= 75).sum() for s in summary_subjects],
            "Distinction in Percent": [round((df[s] >= 75).sum() * 100 / len(df), 2) for s in summary_subjects],
            "Obtained 100 out of 100": [(df[s] == 100).sum() for s in summary_subjects],
        }

        ws_summary = wb.create_sheet(title="Summary")
        for i, (label, values) in enumerate(summary.items(), start=2):
            ws_summary.cell(row=i, column=1).value = label
            ws_summary.cell(row=i, column=1).font = Font(bold=True)
            for j, value in enumerate(values, start=2):
                ws_summary.cell(row=i, column=j).value = value
                if label == "Highest in Subject":
                    ws_summary.cell(row=i, column=j).font = Font(color="FF0000")
                elif label in ["Average in Subject", "Distinction in Subject"]:
                    ws_summary.cell(row=i, column=j).font = Font(color="0000FF")

        wb.save(filename)

    def init_final_frame(self):
        self.final_frame = tk.Frame(self)
        self.final_frame.pack(fill='both', expand=True)
        tk.Label(self.final_frame, text="\ud83c\udf89 Thank you for using the CBSE Result Software!", font=("Arial", 16), fg='green').pack(pady=60)
        tk.Button(self.final_frame, text="Exit", command=self.destroy, font=("Tahoma", 12), width=20, bd=0, bg="red", fg="white").pack(pady=10)



if __name__ == "__main__":
    app = CBSEParserApp()
    app.mainloop()
