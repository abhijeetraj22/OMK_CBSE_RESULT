import tkinter as tk
from tkinter import filedialog, messagebox
import re
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font

# -------------------------
# Subject header maps
# -------------------------
GRADE10_SUBJECT_HEADER_MAP = {
    "184": "ENG-184",
    "085": "HND-085",
    "122": "SNK-122",
    "041": "MAT-041",
    "241": "MAT-241",
    "086": "SCI-086",
    "087": "SST-087",
    "402": "IT-402"
}

GRADE12_SUBJECT_HEADER_MAP = {
    "301": "ENG-301",
    "042": "PHY-042",
    "043": "CHE-043",
    "044": "BIO-044",
    "041": "MAT-041",
    "055": "ACC-055",
    "054": "BST-054",
    "030": "ECO-030",
    "048": "PHED-048",
    "065": "IP-065"
}

# -------------------------
# Main subject sets
# -------------------------
MAIN_SUBJECTS_10 = {"MAT-041", "MAT-241", "SCI-086", "SST-087", "ENG-184", "HND-085", "SNK-122"}
MAIN_SUBJECTS_12_PCM = {"ENG-301", "PHY-042", "CHE-043", "MAT-041", "PHED-048", "IP-065"}
MAIN_SUBJECTS_12_PCB = {"ENG-301", "PHY-042", "CHE-043", "BIO-044", "PHED-048", "IP-065"}
MAIN_SUBJECTS_12_COM = {"ENG-301", "ECO-030", "BST-054", "ACC-055", "PHED-048", "IP-065"}

# -------------------------
# Helper
# -------------------------
def grade_col_for(label):
    return f"{label}-G"


# -------------------------
# Main App
# -------------------------
class CBSEParserApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CBSE Result Soft 1.3 - Launcher")
        self.geometry("520x240")
        self.resizable(False, False)
        self.center_window(520, 240)

        tk.Label(self, text="Choose program version", font=("Arial", 16)).pack(pady=12)
        tk.Button(self, text="Grade 10 Program", width=22, height=2, bg="lightblue",
                  command=lambda: self.open_parser(False)).pack(pady=8)
        tk.Button(self, text="Grade 12 Program", width=22, height=2, bg="lightgreen",
                  command=lambda: self.open_parser(True)).pack(pady=8)
        tk.Button(self, text="Exit", width=10, command=self.destroy).pack(pady=10)

        # Icon safety for PyInstaller
        def resource_path(rel_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, rel_path)

        try:
            self.iconbitmap(resource_path("cbse_TNC_icon.ico"))
        except Exception:
            pass

    def center_window(self, w, h):
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x, y = (sw - w) // 2, (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def open_parser(self, is_grade12: bool):        
        parser = ParserWindow(self, is_grade12)
        parser.grab_set()


# -------------------------
# Parser Window
# -------------------------
class ParserWindow(tk.Toplevel):
    def __init__(self, master=None, is_grade12=False):
        super().__init__(master)
        self.title("CBSE Result Soft 1.4")
        self.geometry("680x520")
        self.resizable(False, False)
        self.center_window(680, 520)

        self.file_path = None
        self.sample_line = ""
        self.grade12 = is_grade12
        self.subject_count = 5 if is_grade12 else 6

        self.init_frame1()

        # Icon safe load
        def resource_path(rel_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, rel_path)

        try:
            self.iconbitmap(resource_path("cbse_TNC_icon.ico"))
        except Exception:
            pass

    def center_window(self, w, h):
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x, y = (sw - w) // 2, (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    # -------------------------
    # UI Step 1
    # -------------------------
    def init_frame1(self):
        self.frame1 = tk.Frame(self)
        self.frame1.pack(fill="both", expand=True, padx=12, pady=12)

        grade_text = "Grade 12" if self.grade12 else "Grade 10"
        tk.Label(self.frame1, text=f"{grade_text} - Step 1: Upload TXT File", font=("Arial", 14)).pack(pady=6)

        btn_frame = tk.Frame(self.frame1)
        btn_frame.pack(pady=6)
        tk.Button(btn_frame, text="Choose File", command=self.select_file,
                  font=("Tahoma", 12), width=15, bd=0, bg="pink").pack(side="left", padx=8)

        self.file_label = tk.Label(btn_frame, text="No file selected", font=("Arial", 10))
        self.file_label.pack(side="left")

        tk.Label(self.frame1, text="Step 2: Paste Sample Student Line", font=("Arial", 12)).pack(pady=8)
        self.sample_entry = tk.Text(self.frame1, height=6, width=80)
        self.sample_entry.pack()

        nav = tk.Frame(self.frame1)
        nav.pack(pady=12)
        tk.Button(nav, text="Back to Launcher", command=self.on_close, width=14).pack(side="left", padx=6)
        tk.Button(nav, text="Next", command=self.handle_sample,
                  font=("Tahoma", 12), width=12, bd=0, bg="lightgreen").pack(side="left", padx=6)

    def on_close(self):
        self.destroy()

    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
        if path:
            self.file_path = path
            self.file_label.config(text=os.path.basename(path))

    def handle_sample(self):
        self.sample_line = self.sample_entry.get("1.0", tk.END).strip()
        if not self.file_path or not self.sample_line:
            messagebox.showerror("Error", "Please select a file and enter a sample line.")
            return

        # --- Determine which grade is selected ---
        if self.grade12:
            expected_codes = set(GRADE12_SUBJECT_HEADER_MAP.keys())
            wrong_codes = set(GRADE10_SUBJECT_HEADER_MAP.keys())
            grade_text = "Grade 12"
        else:
            expected_codes = set(GRADE10_SUBJECT_HEADER_MAP.keys())
            wrong_codes = set(GRADE12_SUBJECT_HEADER_MAP.keys())
            grade_text = "Grade 10"

        # --- Read uploaded file briefly to check codes ---
        with open(self.file_path, "r", encoding="utf-8", errors="ignore") as f:
            file_text = f.read()

        found_codes_in_file = set(re.findall(r"\b\d{3}\b", file_text))
        found_codes_in_sample = set(re.findall(r"\b\d{3}\b", self.sample_line))

        # --- Detect mismatches ---
        wrong_in_file = found_codes_in_file & wrong_codes
        wrong_in_sample = found_codes_in_sample & wrong_codes
        found_expected_in_file = bool(found_codes_in_file & expected_codes)
        found_expected_in_sample = bool(found_codes_in_sample & expected_codes)

        # --- Validation ---
        if not (found_expected_in_file and found_expected_in_sample) or wrong_in_file or wrong_in_sample:
            error_msg = f"Please upload and insert the correct file and sample text for {grade_text}.\n\n"
            if wrong_in_file:
                error_msg += f"❌ Wrong subject codes found in file: {', '.join(sorted(wrong_in_file))}\n"
            if wrong_in_sample:
                error_msg += f"❌ Wrong subject codes found in sample text: {', '.join(sorted(wrong_in_sample))}\n"
            if not found_expected_in_file:
                error_msg += "⚠️ No valid subject codes found in the uploaded file.\n"
            if not found_expected_in_sample:
                error_msg += "⚠️ No valid subject codes found in the sample text.\n"

            messagebox.showerror("Invalid File or Text", error_msg)
            return

        # --- If all good, move to next step ---
        self.frame1.destroy()
        self.init_frame2()



    # -------------------------
    # UI Step 2
    # -------------------------
    def init_frame2(self):
        self.frame2 = tk.Frame(self)
        self.frame2.pack(fill="both", expand=True, padx=12, pady=12)
        tk.Label(self.frame2, text="Step 3: Save Excel Output", font=("Arial", 14)).pack(pady=24)
        tk.Button(self.frame2, text="Save and Generate Excel", command=self.generate_excel,
                  font=("Tahoma", 12), width=28, bd=0, bg="lightblue").pack(pady=8)
        tk.Button(self.frame2, text="Back", command=self.back_to_step1, width=12).pack(pady=6)

    def back_to_step1(self):
        self.frame2.destroy()
        self.init_frame1()

    def generate_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            return
        try:
            self.parse_and_save(file_path)
            response = messagebox.askyesno("Success", f"\u2705 File saved:\n{file_path}\n\nGenerate another file?")
            if response:
                self.frame2.destroy()
                self.init_frame1()
            else:
                self.frame2.destroy()
                self.init_final_frame()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # -------------------------
    # Core Parser
    # -------------------------
    def parse_and_save(self, filename):
        with open(self.file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = [ln.strip() for ln in f if ln.strip()]

        data = []
        i = 0
        while i < len(lines) - 1:
            line1 = lines[i]
            line2 = lines[i + 1] if i + 1 < len(lines) else ""

            if not re.match(r"^\d{8}\s", line1):
                i += 1
                continue

            m = re.match(r"^(\d{8})\s+([MF])\s+(.*?)\s+(\d{3})", line1)
            if not m:
                i += 1
                continue

            roll, gender, name = m.group(1), m.group(2), m.group(3).strip()
            codes = re.findall(r"\b\d{3}\b", line1)
            marks = re.findall(r"(\d{1,3})\s+[A-D][12]", line2)
            marks = [int(m) for m in marks]

            header_map = GRADE12_SUBJECT_HEADER_MAP if self.grade12 else GRADE10_SUBJECT_HEADER_MAP
            row = {"Roll No": roll, "Gender": gender, "Name": name}

            for j in range(min(len(codes), len(marks))):
                code = codes[j]
                label = header_map.get(code, f"SUB-{code}")
                row[label] = marks[j]

            result = re.search(r"\b(PASS|FAIL|COMP|ABST)\b", line1 + " " + line2, re.I)
            row["Result"] = result.group(1).upper() if result else ""

            # Calculations
            if self.grade12:
                vals = [v for k, v in row.items() if k in GRADE12_SUBJECT_HEADER_MAP.values()]
                total = sum(vals)
                row["Total"] = total
                row["Percentage"] = round(total / len(vals), 2) if vals else 0
            else:
                vals = [v for k, v in row.items() if k in GRADE10_SUBJECT_HEADER_MAP.values()]
                main_scores = [v for k, v in row.items() if k in MAIN_SUBJECTS_10]
                top5 = sorted(vals, reverse=True)[:5]
                row["Main Total"] = sum(main_scores)
                row["Main %"] = round(sum(main_scores) / len(main_scores), 2) if main_scores else 0
                row["Top 5 Total"] = sum(top5)
                row["Top 5 %"] = round(sum(top5) / 5, 2) if top5 else 0

            data.append(row)
            i += 2

        df = pd.DataFrame(data)

        # Rank Columns
        if self.grade12:
            if "Percentage" in df.columns:
                df["Rank"] = df["Percentage"].rank(method="dense", ascending=False).astype(int)
        else:
            if "Main %" in df.columns:
                df["Main % Rank"] = df["Main %"].rank(method="dense", ascending=False).astype(int)
            if "Top 5 %" in df.columns:
                df["Top 5 % Rank"] = df["Top 5 %"].rank(method="dense", ascending=False).astype(int)

        # Final column selection
        if self.grade12:
            final_cols = [
                "Roll No", "Gender", "Name",
                "ACC-055", "BIO-044", "BST-054", "CHE-043", "ECO-030",
                "ENG-301", "IP-065", "MAT-041", "PHED-048", "PHY-042",
                "Result", "Total", "Percentage", "Rank"
            ]
        else:
            final_cols = [
                "Roll No", "Gender", "Name",
                "ENG-184", "HND-085", "IT-402", "MAT-041", "MAT-241",
                "SCI-086", "SNK-122", "SST-087",
                "Result", "Main Total", "Main %", "Main % Rank",
                "Top 5 Total", "Top 5 %", "Top 5 % Rank"
            ]

        df = df.reindex(columns=[c for c in final_cols if c in df.columns])
        df.to_excel(filename, index=False)

        # -----------------------------------------
        # 🧾 Add Summary Sheet (new feature)
        # -----------------------------------------
        from openpyxl import load_workbook
        from openpyxl.styles import Font

        # Automatically identify subject columns
        ordered_mark_cols = [col for col in df.columns if "-" in col and df[col].dtype != "object"]

        # If no subject columns found, skip summary
        if ordered_mark_cols:
            wb = load_workbook(filename)
            ws = wb.active

            summary = {
                "Subject": ordered_mark_cols,
                "Highest in Subject": [df[s].max() for s in ordered_mark_cols],
                "Lowest in Subject": [df[s].min() for s in ordered_mark_cols],
                "Average in Subject": [round(df[s].mean(), 2) for s in ordered_mark_cols],
                "Distinction in Subject": [(df[s] >= 75).sum() for s in ordered_mark_cols],
                "Distinction in Percent": [round((df[s] >= 75).sum() * 100 / len(df), 2) for s in ordered_mark_cols],
                "Obtained 100 out of 100": [(df[s] == 100).sum() for s in ordered_mark_cols],
            }

            ws_summary = wb.create_sheet(title="Summary")

            # Write summary data
            for i, (label, values) in enumerate(summary.items(), start=2):
                ws_summary.cell(row=i, column=1).value = label
                ws_summary.cell(row=i, column=1).font = Font(bold=True)
                for j, value in enumerate(values, start=2):
                    ws_summary.cell(row=i, column=j).value = value
                    if label == "Highest in Subject":
                        ws_summary.cell(row=i, column=j).font = Font(color="FF0000")
                    elif label in ["Average in Subject", "Distinction in Subject"]:
                        ws_summary.cell(row=i, column=j).font = Font(color="0000FF")

            wb.save(filename)


    # -------------------------
    # End Frame
    # -------------------------
    def init_final_frame(self):
        self.final_frame = tk.Frame(self)
        self.final_frame.pack(fill="both", expand=True, padx=12, pady=12)
        tk.Label(self.final_frame, text="🎉 Thank you for using CBSE Result Software!", font=("Arial", 16), fg="green").pack(pady=40)
        tk.Button(self.final_frame, text="Exit", command=self.destroy, font=("Tahoma", 12), width=20, bd=0, bg="red", fg="white").pack(pady=10)


if __name__ == "__main__":
    app = CBSEParserApp()
    app.mainloop()
