import tkinter as tk
from tkinter import filedialog, messagebox
import re
import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.styles import Font

# -------------------------
# Subject header maps
# -------------------------
GRADE10_SUBJECT_HEADER_MAP = {
    "184": "ENG-184",
    "085": "HND-085",
    "122": "SNK-122",
    "041": "MAT-041",
    "241": "MAT-241",
    "086": "SCI-086",
    "087": "SST-087",
    "402": "IT-402"
}

GRADE12_SUBJECT_HEADER_MAP = {
    "301": "ENG-301",
    "042": "PHY-042",
    "043": "CHE-043",
    "044": "BIO-044",
    "041": "MAT-041",
    "055": "ACC-055",
    "054": "BST-054",
    "030": "ECO-030",
    "048": "PHED-048",
    "065": "IP-065",
    "027": "HIS-027",
    "028": "POL_SC-028",
    "029": "GEO-029",
    "049": "PAINT-049"
}

# -------------------------
# Main subject sets
# -------------------------
MAIN_SUBJECTS_10 = {"MAT-041", "MAT-241", "SCI-086", "SST-087", "ENG-184", "HND-085", "SNK-122"}
MAIN_SUBJECTS_12_PCM = {"ENG-301", "PHY-042", "CHE-043", "MAT-041", "PHED-048", "IP-065"}
MAIN_SUBJECTS_12_PCB = {"ENG-301", "PHY-042", "CHE-043", "BIO-044", "PHED-048", "IP-065"}
MAIN_SUBJECTS_12_COM = {"ENG-301", "ECO-030", "BST-054", "ACC-055", "PHED-048", "IP-065"}
MAIN_SUBJECTS_12_ARTS = {"ENG-301", "HIS-027", "POL_SC-028", "GEO-029", "PAINT-049", "IP-065"}


# -------------------------
# Main App
# -------------------------
class CBSEParserApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CBSE Result Soft 1.4 - Launcher")
        self.geometry("520x240")
        self.resizable(False, False)
        self.center_window(520, 240)
        self.protocol("WM_DELETE_WINDOW", self.init_final_frame)

        tk.Label(self, text="Choose program version", font=("Arial", 16)).pack(pady=12)
        tk.Button(self, text="Grade 10 Program", width=22, height=2, bg="lightblue",
                  command=lambda: self.open_parser(False)).pack(pady=8)
        tk.Button(self, text="Grade 12 Program", width=22, height=2, bg="lightgreen",
                  command=lambda: self.open_parser(True)).pack(pady=8)
        tk.Button(self, text="Exit", width=10,height=2, bg="red", command=self.init_final_frame).pack(pady=10)

        # Try to load app icon
        def resource_path(rel_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, rel_path)
        try:
            self.iconbitmap(resource_path("cbse_TNC_icon.ico"))
        except Exception:
            pass

    def center_window(self, w, h):
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x, y = (sw - w) // 2, (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    def open_parser(self, is_grade12: bool):
        self.withdraw()
        parser = ParserWindow(self, is_grade12)
        parser.grab_set()
    
    # ✅ Add this inside the launcher
    def init_final_frame(self):
        # Clear launcher widgets
        for widget in self.winfo_children():
            widget.destroy()

        tk.Label(self, text="🎉 Thank you for using CBSE Result Software!",
                 font=("Arial", 16), fg="green").pack(pady=40)

        tk.Button(self, text="Close App", command=self.destroy,
                  font=("Tahoma", 12), width=20, bd=0, bg="red", fg="white").pack(pady=10)


# -------------------------
# Parser Window
# -------------------------
class ParserWindow(tk.Toplevel):
    def __init__(self, master=None, is_grade12=False):
        super().__init__(master)
        self.master = master
        self.grade12 = is_grade12
        self.title("CBSE Result Soft 1.4")
        self.geometry("680x520")
        self.resizable(False, False)
        self.center_window(680, 520)

        # -------------------------
        # Handle "X" button safely
        # -------------------------
        self.protocol("WM_DELETE_WINDOW", self.on_close)

        self.file_path = None
        self.sample_line = ""

        self.init_frame1()

        def resource_path(rel_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            return os.path.join(base_path, rel_path)

        try:
            self.iconbitmap(resource_path("cbse_TNC_icon.ico"))
        except Exception:
            pass
        
    def on_close(self):
        """Triggered when user clicks window close (X) button."""
        try:
            self.destroy()          # Fully destroy this parser window
            self.master.deiconify() # Restore launcher window
        except Exception as e:
            print(f"Error closing window: {e}")

    def center_window(self, w, h):
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        x, y = (sw - w) // 2, (sh - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    # -------------------------
    # Frame 1: Upload + Validate
    # -------------------------
    def init_frame1(self):
        self.frame1 = tk.Frame(self)
        self.frame1.pack(fill="both", expand=True, padx=12, pady=12)

        grade_text = "Grade 12" if self.grade12 else "Grade 10"
        tk.Label(self.frame1, text=f"{grade_text} - Step 1: Upload TXT File", font=("Arial", 14)).pack(pady=6)

        btn_frame = tk.Frame(self.frame1)
        btn_frame.pack(pady=6)
        tk.Button(btn_frame, text="Choose File", command=self.select_file,
                  font=("Tahoma", 12), width=15, bd=0, bg="pink").pack(side="left", padx=8)
        self.file_label = tk.Label(btn_frame, text="No file selected", font=("Arial", 10))
        self.file_label.pack(side="left")

        tk.Label(self.frame1, text="Step 2: Paste Sample Student Line", font=("Arial", 12)).pack(pady=8)
        self.sample_entry = tk.Text(self.frame1, height=6, width=80)
        self.sample_entry.pack()

        nav = tk.Frame(self.frame1)
        nav.pack(pady=12)
        tk.Button(nav, text="Back to Launcher", command=self.back_to_launcher, width=14).pack(side="left", padx=6)
        tk.Button(nav, text="Validate & Next →", command=self.validate_sample,
                  font=("Tahoma", 12), width=16, bd=0, bg="lightgreen").pack(side="left", padx=6)

    def back_to_launcher(self):
        self.destroy()
        self.master.deiconify()

    def select_file(self):
        path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
        if path:
            self.file_path = path
            self.file_label.config(text=os.path.basename(path))

    def validate_sample(self):
        self.sample_line = self.sample_entry.get("1.0", tk.END).strip()
        if not self.file_path or not self.sample_line:
            messagebox.showerror("Error", "Please select a file and enter a sample line.")
            return

        # Determine selected grade
        if self.grade12:
            valid_codes = set(GRADE12_SUBJECT_HEADER_MAP.keys())
            other_codes = set(GRADE10_SUBJECT_HEADER_MAP.keys())
            grade_text = "Grade 12"
        else:
            valid_codes = set(GRADE10_SUBJECT_HEADER_MAP.keys())
            other_codes = set(GRADE12_SUBJECT_HEADER_MAP.keys())
            grade_text = "Grade 10"

        # Common codes (e.g., 041 appears in both)
        common_codes = set(GRADE10_SUBJECT_HEADER_MAP.keys()) & set(GRADE12_SUBJECT_HEADER_MAP.keys())

        # --- Read only the first student line ---
        with open(self.file_path, "r", encoding="utf-8", errors="ignore") as f:
            first_line = ""
            for line in f:
                if re.match(r"^\d{8}\s", line.strip()):
                    first_line = line.strip()
                    break

        if not first_line:
            messagebox.showerror("Error", "No valid student record found in the file.")
            return

        found_codes_in_file = set(re.findall(r"\b\d{3}\b", first_line))
        found_codes_in_sample = set(re.findall(r"\b\d{3}\b", self.sample_line))

        # Ignore common codes when checking for wrong ones
        wrong_in_file = (found_codes_in_file & other_codes) - common_codes
        wrong_in_sample = (found_codes_in_sample & other_codes) - common_codes

        found_expected_in_file = bool(found_codes_in_file & valid_codes)
        found_expected_in_sample = bool(found_codes_in_sample & valid_codes)

        # --- Validation ---
        if not (found_expected_in_file and found_expected_in_sample) or wrong_in_file or wrong_in_sample:
            msg = f"Please upload the correct file and sample for {grade_text}.\n\n"
            if wrong_in_file:
                msg += f"❌ Wrong subject codes in first student record: {', '.join(sorted(wrong_in_file))}\n"
            if wrong_in_sample:
                msg += f"❌ Wrong subject codes in sample text: {', '.join(sorted(wrong_in_sample))}\n"
            if not found_expected_in_file:
                msg += "⚠️ No valid subject codes found in first student record.\n"
            if not found_expected_in_sample:
                msg += "⚠️ No valid subject codes found in sample text.\n"
            messagebox.showerror("Invalid File or Sample", msg)
            return

        messagebox.showinfo("Success", f"✅ File and sample text validated successfully for {grade_text}.")
        self.goto_next_frame()


    def goto_next_frame(self):
        self.frame1.destroy()
        self.init_frame2()

    # -------------------------
    # Frame 2: Save Excel
    # -------------------------
    def init_frame2(self):
        self.frame2 = tk.Frame(self)
        self.frame2.pack(fill="both", expand=True, padx=12, pady=12)
        tk.Label(self.frame2, text="Step 3: Save Excel Output", font=("Arial", 14)).pack(pady=24)
        tk.Button(self.frame2, text="Save and Generate Excel", command=self.generate_excel,
                  font=("Tahoma", 12), width=28, bd=0, bg="lightblue").pack(pady=8)
        tk.Button(self.frame2, text="Back", command=self.back_to_step1, width=12).pack(pady=6)

    def back_to_step1(self):
        self.frame2.destroy()
        self.init_frame1()

    # -------------------------
    # Excel generation
    # -------------------------
    def generate_excel(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not file_path:
            return
        try:
            self.parse_and_save(file_path)
            response = messagebox.askyesno("Success", f"✅ File saved:\n{file_path}\n\nGenerate another file?")
            if response:
                self.frame2.destroy()
                self.init_frame1()
            else:
                self.frame2.destroy()
                self.init_final_frame()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # -------------------------
    # Parser Logic
    # -------------------------
    def parse_and_save(self, filename):
        with open(self.file_path, "r", encoding="utf-8", errors="ignore") as f:
            lines = [ln.strip() for ln in f if ln.strip()]

        data = []
        i = 0
        while i < len(lines) - 1:
            line1 = lines[i]
            line2 = lines[i + 1] if i + 1 < len(lines) else ""

            if not re.match(r"^\d{8}\s", line1):
                i += 1
                continue

            m = re.match(r"^(\d{8})\s+([MF])\s+(.*?)\s+(\d{3})", line1)
            if not m:
                i += 1
                continue

            roll, gender, name = m.group(1), m.group(2), m.group(3).strip()
            codes = re.findall(r"\b\d{3}\b", line1)
            marks = [int(m) for m in re.findall(r"(\d{1,3})\s+[A-D][12]", line2)]

            header_map = GRADE12_SUBJECT_HEADER_MAP if self.grade12 else GRADE10_SUBJECT_HEADER_MAP
            row = {"Roll No": roll, "Gender": gender, "Name": name}

            for j in range(min(len(codes), len(marks))):
                code = codes[j]
                label = header_map.get(code, f"SUB-{code}")
                row[label] = marks[j]

            result = re.search(r"\b(PASS|FAIL|COMP|ABST)\b", line1 + " " + line2, re.I)
            row["Result"] = result.group(1).upper() if result else ""

            # Calculate totals
            if self.grade12:
                vals = [v for k, v in row.items() if k in GRADE12_SUBJECT_HEADER_MAP.values()]
                total = sum(vals)
                row["Total"] = total
                row["Percentage"] = round(total / len(vals), 2) if vals else 0
            else:
                vals = [v for k, v in row.items() if k in GRADE10_SUBJECT_HEADER_MAP.values()]
                main_scores = [v for k, v in row.items() if k in MAIN_SUBJECTS_10]
                top5 = sorted(vals, reverse=True)[:5]
                row["Main Total"] = sum(main_scores)
                row["Main %"] = round(sum(main_scores) / len(main_scores), 2) if main_scores else 0
                row["Top 5 Total"] = sum(top5)
                row["Top 5 %"] = round(sum(top5) / 5, 2) if top5 else 0

            data.append(row)
            i += 2

        df = pd.DataFrame(data)

        if self.grade12:
            if "Percentage" in df.columns:
                df["Rank"] = df["Percentage"].rank(method="dense", ascending=False).astype(int)
        else:
            if "Main %" in df.columns:
                df["Main % Rank"] = df["Main %"].rank(method="dense", ascending=False).astype(int)
            if "Top 5 %" in df.columns:
                df["Top 5 % Rank"] = df["Top 5 %"].rank(method="dense", ascending=False).astype(int)

                # ✅ Reorder columns properly before saving
        if self.grade12:
            final_cols = [
                "Roll No", "Gender", "Name",
                "ENG-301", "PHY-042", "CHE-043", "MAT-041", "BIO-044",
                "ECO-030", "BST-054", "ACC-055", "HIS-027", "POL_SC-028",
                "GEO-029", "PAINT-049", "PHED-048", "IP-065",
                "Result", "Total", "Percentage", "Rank"
            ]
        else:
            final_cols = [
                "Roll No", "Gender", "Name",
                "ENG-184", "HND-085", "IT-402", "MAT-041", "MAT-241",
                "SCI-086", "SNK-122", "SST-087",
                "Result", "Main Total", "Main %", "Main % Rank",
                "Top 5 Total", "Top 5 %", "Top 5 % Rank"
            ]

        # Keep only columns that exist
        df = df.reindex(columns=[c for c in final_cols if c in df.columns])

        df.to_excel(filename, index=False)


        # ✅ Summary Sheet
        ordered_mark_cols = [col for col in df.columns if "-" in col and df[col].dtype != "object"]
        if ordered_mark_cols:
            wb = load_workbook(filename)
            ws_summary = wb.create_sheet(title="Summary")

            summary = {
                "Subject": ordered_mark_cols,
                "Highest": [df[s].max() for s in ordered_mark_cols],
                "Lowest": [df[s].min() for s in ordered_mark_cols],
                "Average": [round(df[s].mean(), 2) for s in ordered_mark_cols],
                "Distinction (≥75)": [(df[s] >= 75).sum() for s in ordered_mark_cols],
                "100 out of 100": [(df[s] == 100).sum() for s in ordered_mark_cols],
            }

            for i, (label, values) in enumerate(summary.items(), start=2):
                ws_summary.cell(row=i, column=1).value = label
                ws_summary.cell(row=i, column=1).font = Font(bold=True)
                for j, value in enumerate(values, start=2):
                    ws_summary.cell(row=i, column=j).value = value

            wb.save(filename)

    def init_final_frame(self):
        self.final_frame = tk.Frame(self)
        self.final_frame.pack(fill="both", expand=True, padx=12, pady=12)
        tk.Label(self.final_frame, text="🎉 Thank you for using CBSE Result Software!",
                 font=("Arial", 16), fg="green").pack(pady=40)
        tk.Button(self.final_frame, text="Exit", command=self.destroy,
                  font=("Tahoma", 12), width=20, bd=0, bg="red", fg="white").pack(pady=10)
        self.master.deiconify()


# -------------------------
# Run App
# -------------------------
if __name__ == "__main__":
    app = CBSEParserApp()
    app.mainloop()
